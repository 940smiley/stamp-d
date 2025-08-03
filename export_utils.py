"""Export utilities for Stamp'd.

Provides helpers to export the database contents to CSV, XLSX and PDF
files.  Exports are written to the ``BACKUP_DIR`` configured in
:mod:`config` and the absolute path of the created file is returned.
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import List

from fpdf import FPDF
from openpyxl import Workbook

from db_utils import Session, Stamp
from config import BACKUP_DIR


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def export_csv() -> str:
    """Export all stamps to a CSV file and return the path."""
    import csv

    session = Session()
    stamps: List[Stamp] = session.query(Stamp).all()
    filepath = os.path.join(BACKUP_DIR, f"export_{_timestamp()}.csv")
    fields = [
        "id",
        "image_path",
        "stamp_name",
        "country",
        "denomination",
        "description",
    ]
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(fields)
        for s in stamps:
            writer.writerow([getattr(s, f) for f in fields])
    return filepath


def export_xlsx() -> str:
    """Export all stamps to an XLSX file and return the path."""
    session = Session()
    stamps: List[Stamp] = session.query(Stamp).all()
    wb = Workbook()
    ws = wb.active
    headers = [
        "ID",
        "Image Path",
        "Name",
        "Country",
        "Denomination",
        "Description",
    ]
    ws.append(headers)
    for s in stamps:
        ws.append(
            [
                s.id,
                s.image_path,
                s.stamp_name,
                s.country,
                s.denomination,
                s.description,
            ]
        )
    filepath = os.path.join(BACKUP_DIR, f"export_{_timestamp()}.xlsx")
    wb.save(filepath)
    return filepath


def export_pdf() -> str:
    """Create a simple PDF catalogue with images and metadata."""
    session = Session()
    stamps: List[Stamp] = session.query(Stamp).all()
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    for s in stamps:
        pdf.add_page()
        if s.image_path and os.path.exists(s.image_path):
            pdf.image(s.image_path, w=60)
        pdf.ln(65)
        pdf.set_font("Arial", size=12)
        pdf.cell(
            0, 10, f"{s.stamp_name or 'Unknown'} - {s.country}", ln=1
        )
        pdf.cell(0, 10, f"Denomination: {s.denomination or ''}", ln=1)
        pdf.multi_cell(0, 10, s.description or "", align="L")
    filepath = os.path.join(BACKUP_DIR, f"export_{_timestamp()}.pdf")
    pdf.output(filepath)
    return filepath
